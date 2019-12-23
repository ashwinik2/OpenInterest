## openintmi.py calculates money inv in call/put
## in each strike price at all requested exp dates

import csv
import pandas as pd
from pandas import DataFrame
from datetime import date
from datetime import datetime
import time
import os
import os.path
import datetime
from os import path
import openpyxl
from datetime import timedelta
from openpyxl.chart import BarChart,Reference,Series,LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.label import DataLabel
import sys
import chartapi
import commonapi

csv.field_size_limit(sys.maxsize)

contractType =['CALL','PUT']
ifilepath = './../datacolls/output/'
ifilestocklist = './../datacolls/input/'+'StockList.csv'

ofilepath = './output/'
ifile = ofilepath
ofilename = ofilepath+'OImax_pain_mp'

money_inv_row_header = ["CallMI", "PutMI","TotalMI"]
TTM =["CallTM","PutTM"]
expdate_chart_y_axis=['MONEY_INV_CALL x K','MONEY_INV_PUT x K','MONEY_INV_TOTAL x K']
OITM =['CallITM','CallOTM','PutITM','PutOTM']

final_chart_col_loc = ['2','12','22','32','42','52','62','72']
expdate_chart_col_loc =['301','401','501']
final_OTM_chart_col_loc = ['2','12','22','32','42','52','62','72']
expdate_chart_insert_pos =["KN24","OJ24","SF24"]

debug = 0
OI_INDEX = 0
SP_INDEX = 1
OP_INDEX = 3

#### Initialize the local lists 

def initLocalList(totoalOptionTypeMoneyInv,OTM,finalOptionExpDateList,getNumberColsData):
    for i in range(0,len(finalOptionExpDateList)):
        totoalOptionTypeMoneyInv.append([])
        for j in range(0,len(contractType)):
            totoalOptionTypeMoneyInv[i].append([])
            for k in range(0,getNumberColsData):
                totoalOptionTypeMoneyInv[i][j].append([])
    for i in range(0,len(finalOptionExpDateList)):
        for j in range(0,len(contractType)):
            for k in range(0,getNumberColsData):
                val = (int)(0)
                totoalOptionTypeMoneyInv[i][j][k] = val

    for i in range(0,len(finalOptionExpDateList)):
        OTM.append([])                
        for j in range(0,getNumberColsData):
            OTM[i].append([])
            for k in range(len(OITM)):
                OTM[i][j].append([])
    for i in range(0,len(finalOptionExpDateList)):
        for j in range(0,getNumberColsData):
            for k in range(len(OITM)):
                val = (int)(0)
                OTM[i][j][k] = val
                
#### Parsing the row cell data

def parseInputRowData(colFrom,data_frame,row):
    rowData = []
    rowData = data_frame.iloc[row,colFrom]
    if(rowData == 'U'):
        OI =0
        SP =0
        OP = 0        
    else:
        rowData = str(rowData).replace('[','')
        rowData = str(rowData).replace(']','')
        rowData = str(rowData).split(':')
        #print("data is :",rowdata)
        OIref =rowData[OI_INDEX]
        if((OIref == 'U') or (OIref == '.') or (OIref == '')):
            OI = (int)(0)
        else:
            OI =(float)(rowData[OI_INDEX])
        OPref =rowData[OP_INDEX]
        if((OPref == 'U') or (OPref =='.') or (OPref == '')):
            OP = (int)(0)
        else:
            OP =(float)(rowData[OP_INDEX])

        SPref =rowData[SP_INDEX]
        if((SPref == 'U') or (SPref =='.') or (SPref == '')):
            SP = (int)(0)
        else:
            SP =(float)(rowData[SP_INDEX])
    return OI,OP,SP

#### Adding call/put money invested of individual strikeprice to excel sheet

def writeMoneyInvDataToXl(contracttype,sheet_obj,strikeprices,strike_money_inv,row,columnList,totoalOptionTypeMoneyInv,expdate):
    
    if(contracttype == commonapi.Contracttype['CALL'].value):
        if(row == 1):
            for i in range(len(money_inv_row_header)):
                col_index = (int)(expdate_chart_col_loc[i])-1
                cellref=sheet_obj.cell(row=1, column=col_index)
                cellref.value=money_inv_row_header[i]

                col_index = (int)(expdate_chart_col_loc[i])
                for item in range(len(strikeprices)):
                    cellref=sheet_obj.cell(row = 1, column=col_index+item)
                    cellref.value=strikeprices[item]

    
    min_col = (int)(expdate_chart_col_loc[contracttype])-1
    print(sheet_obj.max_column)
    
    current_row_index = row
    cellref=sheet_obj.cell(row =current_row_index+1, column=min_col)
    cellref.value=columnList
    col_index = (int)(expdate_chart_col_loc[contracttype])
    value = 0
    for item in range(0,len(strike_money_inv)):                            
        cellref=sheet_obj.cell(row = current_row_index+1, column=col_index+item)
        cellref.value=strike_money_inv[item]
        value += strike_money_inv[item]
    totoalOptionTypeMoneyInv[expdate][contracttype][row-1] = value

    return totoalOptionTypeMoneyInv

#### Adding total call/put money invested of individual strikeprice to excel sheet

def writeTotalMoneyInvDatatoXl(sheet_obj,strikeprices):
        max_rows = sheet_obj.max_row            
        min_col = (int)(expdate_chart_col_loc[2])-1

        cellref=sheet_obj.cell(row=1, column=min_col)
        cellref.value=money_inv_row_header[2]
        
        for item in range(len(strikeprices)):
            cellref=sheet_obj.cell(row = 1, column=min_col+item+1)
            cellref.value=strikeprices[item]

        for row in range(2, max_rows+1):
            cellref=sheet_obj.cell(row = row, column=(int)(expdate_chart_col_loc[0])-1)
            value = cellref.value
            cellref=sheet_obj.cell(row = row, column=min_col)
            cellref.value=value
            
            callvalue = []
            putvalue = []

            col_index = (int)(expdate_chart_col_loc[0])
            for item in range(len(strikeprices)):
                cellref=sheet_obj.cell(row = row, column=col_index+item)
                value = cellref.value
                callvalue.append(value)                
                
            col_index = (int)(expdate_chart_col_loc[1])   
            for item in range(len(strikeprices)):
                cellref=sheet_obj.cell(row = row, column=col_index+item)
                value = cellref.value
                putvalue.append(value)
            
            for item in range(len(strikeprices)):
                cellref=sheet_obj.cell(row = row, column=min_col+item+1)                        
                total = callvalue[item]+putvalue[item]
                cellref.value=total

#Adding last sheet "ALL_MICP" to workbook which contain data like OTM,ITM,total money inv of call/put of requested expiration dates
def addSheetToXl(wb_obj,symbol):
    
    if symbol in wb_obj.sheetnames:
        print("wb_obj.sheetnames is :",wb_obj.sheetnames)
        std=wb_obj.get_sheet_by_name(symbol)
        wb_obj.remove_sheet(std)
    if not symbol in wb_obj.sheetnames:
        wb_obj.create_sheet(symbol)

#writing call/put total money inv to last sheet "ALL_MICP"
def writeCPMoneyInvToCPMISheet(sheet_obj,number_of_expiration,totoalOptionTypeMoneyInv,getNumberColsData,datesList):

    cellref = sheet_obj.cell(row = (int)(final_chart_col_loc[0])-1,column = 1)
    cellref.value = "MICP"
    
    for item in range(len(datesList)):
        cellref = sheet_obj.cell(row = item+(int)(final_chart_col_loc[0]),column = 1)
        cellref.value = datesList[item]
       
    for i in range(0,number_of_expiration):
        min_cols = (int)(final_chart_col_loc[i])
        max_cols =min_cols+1
        for j in range(len(TTM)):
            min_cols += j
            cellref = sheet_obj.cell(row = 1,column =min_cols)
            cellref.value = TTM[j]            
            
        min_cols = (int)(final_chart_col_loc[i])
        for j in range(len(TTM)):            
            for k in range(0,getNumberColsData): 
                cellref=sheet_obj.cell(row = k+2, column=min_cols)                        
                total = totoalOptionTypeMoneyInv[i][j][k]
                cellref.value=(int)(total)
            min_cols += 1


#writing call/put OTM,ITM inv to last sheet "ALL_MICP"
def writeCPIOTMtoCPMISheet(OTM,sheet_obj,number_of_expiration,getNumberColsData,datesList):
    
    min_rows = 42
    cellref = sheet_obj.cell(row = min_rows,column =1)
    cellref.value = "Dates"

    for item in range(len(datesList)):
        cellref = sheet_obj.cell(row = item+min_rows+1,column = 1)
        cellref.value = datesList[item]

    for i in range(0,number_of_expiration):
        min_cols = (int)(final_OTM_chart_col_loc[i])
        max_cols =min_cols+3
        for j in range(0,4):
            cellref = sheet_obj.cell(row= min_rows,column =min_cols+j)
            cellref.value = OITM[j]
            
        min_cols = (int)(final_OTM_chart_col_loc[i])   
        for row in range(0,getNumberColsData):               
            for col in range(0,4):
                cellref=sheet_obj.cell(row = row+min_rows+1, column=min_cols+col)
                total = OTM[i][row][col]
                cellref.value=(int)(total)


def getStrikePriceMoneyInv(strikepriceList,stockprice,strikePriceFromCSV,row,OptionExpDate,colFrom,data_frame,contracttype,money_inv,col,OTM,expdate):
    strikeexpdate = commonapi.getOptExpDateFromCSV(strikePriceFromCSV[row])
    strikeprice = (float)(commonapi.getOptionStrikePrice(strikePriceFromCSV[row]))
    strikeprice = (float)(strikeprice)

    if((strikeexpdate == OptionExpDate) and ((float)(strikepriceList[stockprice]) == strikeprice)):
        strikePrice = commonapi.getOptionStrikePrice(strikePriceFromCSV[row])                        
        stockPrice = strikeprice
        
        stockPrice = (float)(stockPrice)
        strikePrice = (float)(strikePrice)
        OI,OP,SP= parseInputRowData(colFrom,data_frame,row)

        money_inv = OI*OP*100
        money_inv = money_inv/1000

        if(contracttype == commonapi.Contracttype['CALL'].value):
            if(SP > strikeprice):
                value = OTM[expdate][col][0]
                value += money_inv 
                OTM[expdate][col][0] = value
            else:
                value = OTM[expdate][col][1]
                value += money_inv 
                OTM[expdate][col][1] = value

        if(contracttype == commonapi.Contracttype['PUT'].value):
            if(SP < strikeprice):
                value = OTM[expdate][col][2]
                value += money_inv 
                OTM[expdate][col][2] = value
            else:
                value = OTM[expdate][col][3]
                value += money_inv 
                OTM[expdate][col][3] = value
    return money_inv,OTM

def money_inv(Symbol,number_of_expiration,getNumberColsData):
    symbol = (str)(Symbol)
    sheet_count = 0
    finalOptionExpDateList =[]
    totoalOptionTypeMoneyInv = [[[]]]
    OTM = [[[]]]
    datesList =[]
    
    oxlfile = commonapi.getoFile(symbol,ofilename)    
    wb_obj = openpyxl.load_workbook(oxlfile)
    sheets = wb_obj.sheetnames
        
    istockCSVfile =  commonapi.getStockCSVFile(0,symbol,ifilepath)
    data_frame = pd.read_csv(istockCSVfile, index_col = False)
    finalOptionExpDateList = commonapi.getExpDatesListFromCSV(data_frame,number_of_expiration)
    
    initLocalList(totoalOptionTypeMoneyInv,OTM,finalOptionExpDateList,getNumberColsData)

#Loop around num of expiration x contracttype x getNumberColsData x strikeprices
                    
    for expdate in range(0,len(finalOptionExpDateList)):
        colFrom = 0
        strikeprices = []
        strike_money_inv = []
        col=[]
        sheet_obj_ref = 0        
        
        for contracttype in range(len(contractType)):
            del strikeprices[:]
            del strike_money_inv[:]
            
            if(contracttype == commonapi.Contracttype['CALL'].value):
                 sym = 'All_Exp_CPMI'
                 addSheetToXl(wb_obj,sym)
            sheet_obj = wb_obj[sheets[expdate]]
            sheet_obj_ref = sheet_obj
 
            istockCSVfile =  commonapi.getStockCSVFile(contracttype,symbol,ifilepath)
            if os.path.exists(istockCSVfile):
                print("File Exists")
                data_frame,columnList,colFrom,strikePriceFromCSV = commonapi.getDataFrame(istockCSVfile,getNumberColsData)

                strikepriceList = []
                strikepriceList = commonapi.strikePricesFromCSV(finalOptionExpDateList[expdate],data_frame)
                strikeprices = strikepriceList
                
                for col in range(0,getNumberColsData):                    
                    del strike_money_inv[:]
                    
                    for stockprice in range(0,len(strikepriceList)):
                        money_inv = 0
                        
                        for row in range(0,len(strikePriceFromCSV)):
                            money_inv,OTM = getStrikePriceMoneyInv(strikepriceList,stockprice,strikePriceFromCSV,row,finalOptionExpDateList[expdate],colFrom,data_frame,contracttype,money_inv,col,OTM,expdate)
                                                                       
                        strike_money_inv.append(money_inv)
                        
             #### Adding call max pain of individual strikeprice to excel sheet                            
                    totoalOptionTypeMoneyInv = writeMoneyInvDataToXl(contracttype,sheet_obj,strikeprices,strike_money_inv,col+1,columnList[colFrom],totoalOptionTypeMoneyInv,expdate)
                    if(expdate == 0 and contracttype == commonapi.Contracttype['CALL'].value ):
                        datesList.append(columnList[colFrom])
                    colFrom += 1
                    
        print("len(dates) and dates  is :",len(datesList),datesList)
        #### Calculating Total money invested of call and put then adding data to excel sheet
        writeTotalMoneyInvDatatoXl(sheet_obj_ref,strikeprices)             

        ### call chartapi to insert chart
        chartapi.insertLineChart(sheet_obj_ref,strikeprices,symbol,finalOptionExpDateList[expdate],expdate_chart_y_axis,expdate_chart_col_loc,expdate_chart_insert_pos)

    sheet_obj = wb_obj["All_Exp_CPMI"]
    
    #writing call/put total money inv to last sheet "ALL_MICP"
    writeCPMoneyInvToCPMISheet(sheet_obj,number_of_expiration,totoalOptionTypeMoneyInv,getNumberColsData,datesList)

    #Insert chart from total money call/put data
    chartapi.insertTotalMoneyCPChart(sheet_obj,number_of_expiration,symbol,finalOptionExpDateList,final_chart_col_loc)

    #writing call/put OTM,ITM to last sheet "ALL_MICP"
    writeCPIOTMtoCPMISheet(OTM,sheet_obj,number_of_expiration,getNumberColsData,datesList)

    #Insert chart from OTM,ITM call/put data
    chartapi.insertCPIOMchartToCPMISheet(sheet_obj,symbol,finalOptionExpDateList,number_of_expiration,final_OTM_chart_col_loc,getNumberColsData)
 
    wb_obj.save(oxlfile)
                    
print('done')
