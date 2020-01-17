
import csv
import pandas as pd
from pandas import DataFrame
from datetime import date
from datetime import datetime
import time
import os
import sys
import os.path
import datetime
from os import path
import xlsxwriter
import openpyxl
from datetime import timedelta
import chartapi
import openintmi
import Commonapi
import dataapi
import globalheader
import moneyinvAllExpDate
csv.field_size_limit(sys.maxsize)

contractType =['CALL','PUT']
ifilepath = './../datacolls/output/'
ifilestocklist = './../datacolls/input/'+'StockList.csv'

ofilepath = './output/'
ofilename = ofilepath+'OImax_pain_AllExp'

#thses datas are used by chart insertion and data insertion of the money invested of call and put
maxpain_row_header = ["CallMPx100", "PutMPx100"]
expdate_chart_y_axis=['MAX_PAIN_CALL x 100','MAX_PAIN_PUT x 100 ']
expdate_chart_col_loc =['2','141']
expdate_chart_insert_pos =["A64","CV64"]


# Adding call/put max pain data of individual strikeprice of all expiration dates of all trading days to excel sheet
# calls insertLineChart which is in chartapi.py 
#writeMaxPainDataToXl(contracttype,wb_obj,max_pain_dict,Symbol):
        #input
        #contracttype - call or put
        #wb_obj - xl file
        # max_pain - max pain dict data
        #symbol - stock expiration symbol
        
def writeMaxPainDataToXl(contracttype,wb_obj,max_pain_dict,Symbol):
        if(Commonapi.info == 1):
                logging.info('writeMaxPainDataToXl Started')
        exp_date_list = max_pain_dict.keys()
        sheet_obj = 0
        strikeprice_list = []

        for expdatekey in range(0,len(exp_date_list)):
                if((contracttype == Commonapi.Contracttype['CALL'].value)):
                        addSheetToXl(wb_obj,Symbol,expdatekey,exp_date_list[expdatekey])
                sheets = wb_obj.sheetnames
                sheet_obj = wb_obj[sheets[expdatekey]]
                
                trading_day_list = max_pain_dict[exp_date_list[expdatekey]].keys()
                for trading_day in range(0,len(trading_day_list)):

                    strikeprice_list = max_pain_dict[exp_date_list[expdatekey]][trading_day_list[trading_day]].keys()
                    strikeprice_value_list = max_pain_dict[exp_date_list[expdatekey]][trading_day_list[trading_day]].values()
                    if(contracttype == Commonapi.Contracttype.CALL.value):
                            col_index = (int)(expdate_chart_col_loc[contracttype])
                            cellref=sheet_obj.cell(row=1, column=col_index-1)
                            cellref.value=maxpain_row_header[contracttype]
                            cellref=sheet_obj.cell(row=trading_day+2, column=1)
                            cellref.value=trading_day_list[trading_day]

                            col_index = (int)(expdate_chart_col_loc[contracttype])
                            for i in range(len(strikeprice_list)):
                                    cellref=sheet_obj.cell(row=1, column=col_index+i)
                                    cellref.value=strikeprice_list[i]
                            for item in range(len(strikeprice_list)):
                                    cellref=sheet_obj.cell(row = trading_day+2, column=col_index+item)
                                    cellref.value=strikeprice_value_list[item]

                    else:
                            col_index = (int)(expdate_chart_col_loc[contracttype])
                            cellref=sheet_obj.cell(row=1, column=col_index-1)
                            cellref.value=maxpain_row_header[contracttype]
                            cellref=sheet_obj.cell(row=trading_day+2, column=col_index-1)
                            cellref.value=trading_day_list[trading_day]
                                            
                            for i in range(len(strikeprice_list)):
                                    cellref=sheet_obj.cell(row=1, column=col_index+i)
                                    cellref.value=strikeprice_list[i]

                            for item in range(len(strikeprice_list)):
                                cellref=sheet_obj.cell(row = trading_day+2, column=col_index+item)
                                cellref.value=strikeprice_value_list[item]
                y_axis_label =[]
                y_axis_label.append(expdate_chart_y_axis[contracttype])
                chart_insert_pos = []
                chart_insert_pos.append(expdate_chart_insert_pos[contracttype])
                chart_col_loc=[]
                chart_col_loc.append(expdate_chart_col_loc[contracttype])
                chartapi.insertLineChart(sheet_obj,strikeprice_list,Symbol,exp_date_list[expdatekey],y_axis_label,chart_col_loc,chart_insert_pos,1)               

        if(Commonapi.info == 1):
                logging.info('writeMaxPainDataToXl Ended')
                
#Adding sheet of each expiration date to xl workbook
#addSheetToXl(wb_obj,symbol,sheetIndex,finalOptionExpDateList):
        #input
        #wb_obj - xl file
        #sheetIndex - sheet index
        #symbol - stock expiration symbol
        #finalOptionExpDateList - stock expiration date list
                
def addSheetToXl(wb_obj,symbol,sheetIndex,finalOptionExpDateList):
        if(Commonapi.info == 1):
                logging.info('addSheetToXl Started')
        sheet_name =  str(symbol)+finalOptionExpDateList
        if sheet_name in wb_obj.sheetnames:
                if(Commonapi.debug == 1):
                        logging.debug('%s %s', 'sheet_name present in wb_obj.sheetnames is', sheet_name,wb_obj.sheetnames) 
                    
        if not sheet_name in wb_obj.sheetnames:
                if(Commonapi.debug == 1):
                        logging.debug('%s %s', 'sheet_name not present in wb_obj.sheetnames is', sheet_name,wb_obj.sheetnames)
                wb_obj.create_sheet(index = sheetIndex , title = sheet_name)
        if(Commonapi.info == 1):
                logging.info('addSheetToXl Ended')

#max pain calls getAllMaxPain(symbol,contractype) which is in commonapi.py for each option type
#gets max pain dict for all existing expiration dates and calls local writeMaxPainDataToXl
        #input
        #symbol - stock symbol
def maxpain(Symbol):
        if(Commonapi.info == 1):
                logging.info('maxpain Started')
        sheet_count = 0
        finalOptionExpDateList =[]
        call_error,istockCSVfilename =  dataapi.getStockCSVFile(0,symbol,ifilepath)
        if(call_error == globalheader.Success): 
                oxlfile = Commonapi.createoFile(symbol,ofilename)
                if os.path.exists(oxlfile):
                        logging.warning('%s %s', 'file exists:', oxlfile)
                        os.remove(oxlfile)
                        logging.warning('File Removed!')
            
                if os.path.exists(oxlfile):
                        wb_obj = openpyxl.load_workbook(oxlfile)
                        logging.warning('%s %s', 'file exists:', oxlfile)
                else:
                        logging.warning('%s %s', 'File does not Exists:', oxlfile)
                        wb_obj = openpyxl.Workbook()
  
                for contracttype in range(len(contractType)):                                    
                        call_error,max_pain_dict = Commonapi.getAllMaxPain(Symbol,contracttype)
                        if(call_error == globalheader.Success):
                                if(Commonapi.info == 1):
                                    logging.info('Commonapi.getMaxPain Success')
                                writeMaxPainDataToXl(contracttype,wb_obj,max_pain_dict,Symbol)
                        else:
                                logging.error('%s %d', 'Commonapi.getMaxPain Failed', call_error)
                               return call_error
                wb_obj.save(oxlfile)
                if(Commonapi.info == 1):
                        logging.info('maxpain Ended')
                return call_error

        else:
                logging.error('%s %d', 'max pain Failed', call_error)
                return call_error

        
                           
if __name__ == "__main__":
        if(Commonapi.info == 1):
                logging.info('maxpain main Started')
        symbol = raw_input("Enter symbol :") 
        call_error = maxpain(symbol)
        if(call_error == globalheader.Success):
                if(Commonapi.info == 1):
                    logging.info('%s %d','maxpain Success', call_error)
                moneyinvAllExpDate.money_inv(symbol)
        else:
                logging.error('%s %d', 'maxpain Failed', call_error)

        if(Commonapi.info == 1):
                logging.info('maxpain main Ended')

                    
