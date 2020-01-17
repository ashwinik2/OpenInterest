## openintmi.py calculates money inv in call/put
## in each strike price at all requested exp dates

import csv
import pandas as pd
from pandas import DataFrame
from datetime import date
from datetime import datetime
import time
import os
import os.path
import datetime
from os import path
import openpyxl
from datetime import timedelta
from openpyxl.chart import BarChart,Reference,Series,LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.label import DataLabel
from collections import OrderedDict
import sys
import chartapi
import Commonapi
import dataapi
import globalheader

csv.field_size_limit(sys.maxsize)

contractType =['CALL','PUT']
ifilepath = './../datacolls/output/'

ofilepath = './output/'
ifile = ofilepath
ofilename = ofilepath+'OImax_pain_AllExp'

#thses datas are used by chart insertion and data insertion of the money invested of call and put

money_inv_row_header = ["CallMIxK", "PutMIxK"]
expdate_chart_y_axis=['MONEY_INV_CALL x K','MONEY_INV_PUT x K']
final_chart_insert_pos = ["A58","K58","T58","AD58","AN58","AX58","BH58","BR58","CB58","CK58","CW58","DF58","DP58","EA58","EK58","EU58","FE58"]
final_OTM_chart_insert_pos = ["A124","K124","T124","AD124","AN124","AX124","BH124","BR124","CB124","CK124","CW124","DF124","DP124","EA124","EK124","EU124","FE124"]
final_OTMOI_chart_insert_pos = ["A200","K200","T200","AD200","AN200","AX200","BH200","BR200","CB200","CK200","CW200","DF200","DP200","EA200","EK200","EU200","FE200"]
final_chart_col_loc = ['2']
expdate_chart_col_loc =['301','451']
final_TM_chart_col_loc = ['2','12','22','32','42','52','62','72','82','92','102','112','122','132','142','152','162','172']
expdate_chart_insert_pos =["KN64","OJ64"]
data_col_loc = []
chart_insert_loc=[]
length = 0
num_of_exp = 0
chart_col_loc =  ['2','12','22','32','42','52','62','72','82','92','102','112','122','132','142']


#Adding last sheet "ALL_MICP" to workbook which contain data like OTM,ITM,total money inv of call/put of requested expiration dates
#addSheetToXl(wb_obj,symbol):
        #input
        #wb_obj - xl file
        #symbol - stock expiration symbol

def addSheetToXl(wb_obj,symbol):
    if(Commonapi.info == 1):
        logging.info('addSheetToXl Started')
    if symbol in wb_obj.sheetnames:
        if(Commonapi.debug == 1):
            logging.debug('%s %s', 'sheet_name present in wb_obj.sheetnames is', symbol,wb_obj.sheetnames) 
        std=wb_obj.get_sheet_by_name(symbol)
        wb_obj.remove_sheet(std)
    if not symbol in wb_obj.sheetnames:
        if(Commonapi.debug == 1):
            logging.debug('%s %s', 'sheet_name not present in wb_obj.sheetnames is', symbol,wb_obj.sheetnames)
        wb_obj.create_sheet(symbol)
    if(Commonapi.info == 1):
        logging.info('addSheetToXl Ended')


# Adding call/put max pain data of individual strikeprice of all expiration dates of all trading days to excel sheet
# calls insertLineChart which is in chartapi.py 
#writeMoneyInvDataToXl(contracttype,wb_obj,money_inv_dict,Symbol):
        #input
        #contracttype - call or put
        #wb_obj - xl file
        # money_inv_dict - money inv dict data
        #symbol - stock expiration symbol        
def writeMoneyInvDataToXl(contracttype,wb_obj,money_inv_dict,Symbol):
        if(Commonapi.info == 1):
            logging.info('writeMoneyInvDataToXl Started')
        global num_of_exp
        exp_date_list = money_inv_dict.keys()
        num_of_exp = len(exp_date_list)
        sheet_obj_ref = 0
        strikeprice_list = []
        
        if(contracttype == Commonapi.Contracttype['CALL'].value):
                sym = 'All_Exp_CPMI'
                addSheetToXl(wb_obj,sym)
        

        for expdatekey in range(0,len(exp_date_list)):
                sheets = wb_obj.sheetnames
                sheet_obj = wb_obj[sheets[expdatekey]]
                sheet_obj_ref = sheet_obj

                trading_day_list = money_inv_dict[exp_date_list[expdatekey]].keys()
                for trading_day in range(0,len(trading_day_list)):
                    strikeprice_list = money_inv_dict[exp_date_list[expdatekey]][trading_day_list[trading_day]].keys()
                    strikeprice_value_list = money_inv_dict[exp_date_list[expdatekey]][trading_day_list[trading_day]].values()
                    
                    if(contracttype == Commonapi.Contracttype.CALL.value):
                            col_index = (int)(expdate_chart_col_loc[contracttype])
                            cellref=sheet_obj.cell(row=1, column=col_index-1)
                            cellref.value=money_inv_row_header[contracttype]
                            cellref=sheet_obj.cell(row=trading_day+2, column=col_index-1)
                            cellref.value=trading_day_list[trading_day]

                            col_index = (int)(expdate_chart_col_loc[contracttype])
                            for i in range(len(strikeprice_list)):
                                    cellref=sheet_obj.cell(row=1, column=col_index+i)
                                    cellref.value=strikeprice_list[i]
                            for item in range(len(strikeprice_list)):
                                    cellref=sheet_obj.cell(row = trading_day+2, column=col_index+item)
                                    cellref.value=strikeprice_value_list[item]

                    else:
                            col_index = (int)(expdate_chart_col_loc[contracttype])
                            cellref=sheet_obj.cell(row=1, column=col_index-1)
                            cellref.value=money_inv_row_header[contracttype]
                            cellref=sheet_obj.cell(row=trading_day+2, column=col_index-1)
                            cellref.value=trading_day_list[trading_day]
                                            
                            for i in range(len(strikeprice_list)):
                                    cellref=sheet_obj.cell(row=1, column=col_index+i)
                                    cellref.value=strikeprice_list[i]

                            for item in range(len(strikeprice_list)):
                                cellref=sheet_obj.cell(row = trading_day+2, column=col_index+item)
                                cellref.value=strikeprice_value_list[item]
                y_axis_label =[]
                y_axis_label.append(expdate_chart_y_axis[contracttype])
                chart_insert_pos = []
                chart_insert_pos.append(expdate_chart_insert_pos[contracttype])
                chart_col_loc=[]
                chart_col_loc.append(expdate_chart_col_loc[contracttype])
                chartapi.insertLineChart(sheet_obj_ref,strikeprice_list,Symbol,exp_date_list[expdatekey],y_axis_label,chart_col_loc,chart_insert_pos,1)

        if(Commonapi.info == 1):
            logging.info('writeMoneyInvDataToXl Ended')

# Adding call/put OTM,ITM money inv data  to last sheet "ALL_MICP of individual strikeprice of all expiration dates of all trading days to excel sheet
# calls insertLineChart which is in chartapi.py 
#writeTotalMoneyInvDatatoXl(contracttype,wb_obj,otm_dict,Symbol):
        #input
        #contracttype - call or put
        #wb_obj - xl file
        # otm_dict - OTM,ITM money inv dict data
        #symbol - stock expiration symbol
def writeCPIOTMtoCPMISheet(contracttype,wb_obj,otm_dict,Symbol):
        if(Commonapi.info == 1):
            logging.info('writeCPIOTMtoCPMISheet Started')
        sheet_obj = wb_obj["All_Exp_CPMI"]
        min_rows = sheet_obj.max_row
        min_cols = sheet_obj.max_column

        rows = min_rows+34 
        exp_date_list = otm_dict.keys()
        for expdatekey in range(0,len(exp_date_list)):
                if((contracttype == Commonapi.Contracttype.CALL.value) and (expdatekey == 0)):
                        data_col_loc.append(rows)
                trading_day_list = otm_dict[exp_date_list[expdatekey]].keys()
                for trading_day in range(0,len(trading_day_list)):
                
                    if((contracttype == Commonapi.Contracttype.CALL.value) and (expdatekey == 0)):
                        chart_insert_loc.append(rows+len(trading_day_list)+2)
                    
                    header_list = otm_dict[exp_date_list[expdatekey]][trading_day_list[trading_day]].keys()
                    header_value_list = otm_dict[exp_date_list[expdatekey]][trading_day_list[trading_day]].values()

                    if(contracttype == Commonapi.Contracttype.CALL.value):
                            col_index = (int)(final_TM_chart_col_loc[expdatekey])
                            cellref = sheet_obj.cell(row = data_col_loc[0],column =col_index-1)
                            cellref.value = str(Symbol)+'_'+str(exp_date_list[expdatekey])
                            
                            cellref=sheet_obj.cell(row=data_col_loc[0]+trading_day+1, column=col_index-1)
                            cellref.value=trading_day_list[trading_day]
                                                    
                            for header_index in range(0,len(header_list)):
                                    cellref = sheet_obj.cell(row = data_col_loc[0],column = col_index+header_index)
                                    cellref.value = header_list[header_index]

                            for header_index in range(0,len(header_list)):
                                    cellref = sheet_obj.cell(row = data_col_loc[0]+1+trading_day,column = col_index+header_index)
                                    cellref.value = header_value_list[header_index]
                                                                                    
                    else:
                            col_index = (int)(final_TM_chart_col_loc[expdatekey])
                            cellref = sheet_obj.cell(row = data_col_loc[0],column =col_index-1)
                            cellref.value = str(Symbol)+'_'+str(exp_date_list[expdatekey])
                            
                            cellref=sheet_obj.cell(row=data_col_loc[0]+trading_day+1, column=col_index-1)
                            cellref.value=trading_day_list[trading_day]

                            col_index = (int)(final_TM_chart_col_loc[expdatekey])+2
                            for header_index in range(0,len(header_list)):
                                    cellref = sheet_obj.cell(row = data_col_loc[0],column = col_index+header_index)
                                    cellref.value = header_list[header_index]

                            for header_index in range(0,len(header_list)):
                                    cellref = sheet_obj.cell(row = data_col_loc[0]+1+trading_day,column = col_index+header_index)
                                    cellref.value = header_value_list[header_index]
        if(Commonapi.info == 1):
            logging.info('writeCPIOTMtoCPMISheet Ended')
        

# Adding call/put total money data of individual strikeprice of all expiration dates of all trading days to excel sheet
# calls insertLineChart which is in chartapi.py 
#writeTotalMoneyInvDatatoXl(contracttype,wb_obj,tm_dict,Symbol):
        #input
        #contracttype - call or put
        #wb_obj - xl file
        # tm_dict - total money dict data
        #symbol - stock expiration symbol
def writeTotalMoneyInvDatatoXl(contracttype,wb_obj,tm_dict,Symbol):
        if(Commonapi.info == 1):
            logging.info('writeTotalMoneyInvDatatoXl Started')
        global length
        sheet_obj = wb_obj["All_Exp_CPMI"]
        min_rows = sheet_obj.max_row
        min_cols = sheet_obj.max_column
        
        exp_date_list = tm_dict.keys()
        for expdatekey in range(0,len(exp_date_list)):

                trading_day_list = tm_dict[exp_date_list[expdatekey]].keys()
                for trading_day in range(0,len(trading_day_list)):
                
                    header_list = tm_dict[exp_date_list[expdatekey]][trading_day_list[trading_day]].keys()
                    header_value_list = tm_dict[exp_date_list[expdatekey]][trading_day_list[trading_day]].values()

                    if((contracttype == Commonapi.Contracttype.CALL.value) and (expdatekey == 0)):                        
                        
                        #data_col_loc.append(min_cols)
                        chart_insert_loc.append(len(trading_day_list)+2+1)
                        length = len(trading_day_list)                        

                    if(contracttype == Commonapi.Contracttype.CALL.value):
                            col_index = (int)(final_TM_chart_col_loc[expdatekey])
                            cellref = sheet_obj.cell(row = 1,column =col_index-1)
                            cellref.value = str(Symbol)+'_'+str(exp_date_list[expdatekey])
                            
                            cellref=sheet_obj.cell(row=trading_day+2, column=col_index-1)
                            cellref.value=trading_day_list[trading_day]
                            
                            cellref = sheet_obj.cell(row = 1,column = col_index)
                            cellref.value = header_list[0]
                            
                            for header_index in range(0,len(header_list)):
                                    cellref = sheet_obj.cell(row = trading_day+2,column = col_index+header_index)
                                    cellref.value = header_value_list[header_index]
                                                                                    
                    else:
                            col_index = (int)(final_TM_chart_col_loc[expdatekey])
                            cellref=sheet_obj.cell(row=trading_day+2, column=col_index-1)
                            cellref.value=trading_day_list[trading_day]
                            
                            cellref = sheet_obj.cell(row = 1,column = col_index+1)
                            cellref.value = header_list[0]
                            
                            col_index = (int)(final_TM_chart_col_loc[expdatekey])+1
                            for header_index in range(0,len(header_list)):
                                    cellref = sheet_obj.cell(row = trading_day+2,column = col_index+header_index)
                                    cellref.value = header_value_list[header_index]
        if(Commonapi.info == 1):
            logging.info('writeTotalMoneyInvDatatoXl Ended')

# Adding call/put in the money and out-of-money  open  interest of individual strikeprice of all expiration dates of all trading days to excel sheet
# calls insertLineChart which is in chartapi.py 
#writeCPIOTMIOtoCPMISheet(contracttype,wb_obj,otmoi_dict,Symbol):
        #input
        #contracttype - call or put
        #wb_obj - xl file
        #otmoi_dict - money inv in-the-money and out-of-the-money open  interest dict data
        #symbol - stock expiration symbol
        
def writeCPIOTMIOtoCPMISheet(contracttype,wb_obj,otmoi_dict,Symbol):
        if(Commonapi.info == 1):
            logging.info('writeCPIOTMIOtoCPMISheet Started')
        sheet_obj = wb_obj["All_Exp_CPMI"]
        min_rows = sheet_obj.max_row
        min_cols = sheet_obj.max_column

        rows = min_rows+34 
                
        exp_date_list = otmoi_dict.keys()
        for expdatekey in range(0,len(exp_date_list)):
                if((contracttype == Commonapi.Contracttype.CALL.value) and (expdatekey == 0)):
                        data_col_loc.append(rows)
                trading_day_list = otmoi_dict[exp_date_list[expdatekey]].keys()
                for trading_day in range(0,len(trading_day_list)):
                
                    if((contracttype == Commonapi.Contracttype.CALL.value) and (expdatekey == 0)):
                        chart_insert_loc.append(rows+len(trading_day_list)+2)
                    
                    header_list = otmoi_dict[exp_date_list[expdatekey]][trading_day_list[trading_day]].keys()
                    header_value_list = otmoi_dict[exp_date_list[expdatekey]][trading_day_list[trading_day]].values()
                    
                    if(contracttype == Commonapi.Contracttype.CALL.value):
                            col_index = (int)(final_TM_chart_col_loc[expdatekey])
                            cellref = sheet_obj.cell(row = data_col_loc[1],column =col_index-1)
                            cellref.value = str(Symbol)+'_'+str(exp_date_list[expdatekey])
                            
                            cellref=sheet_obj.cell(row=data_col_loc[1]+trading_day+1, column=col_index-1)
                            cellref.value=trading_day_list[trading_day]
                                                    
                            for header_index in range(0,len(header_list)):
                                    cellref = sheet_obj.cell(row = data_col_loc[1],column = col_index+header_index)
                                    cellref.value = header_list[header_index]

                            for header_index in range(0,len(header_list)):
                                    cellref = sheet_obj.cell(row = data_col_loc[1]+1+trading_day,column = col_index+header_index)
                                    cellref.value = header_value_list[header_index]
                                                                                    
                    else:

                            col_index = (int)(final_TM_chart_col_loc[expdatekey])
                            cellref = sheet_obj.cell(row = data_col_loc[1],column =col_index-1)
                            cellref.value = str(Symbol)+'_'+str(exp_date_list[expdatekey])
                            
                            cellref=sheet_obj.cell(row=data_col_loc[0]+trading_day+1, column=col_index-1)
                            cellref.value=trading_day_list[trading_day]

                            col_index = (int)(final_TM_chart_col_loc[expdatekey])+2
                            for header_index in range(0,len(header_list)):
                                    cellref = sheet_obj.cell(row = data_col_loc[1],column = col_index+header_index)
                                    cellref.value = header_list[header_index]

                            for header_index in range(0,len(header_list)):
                                    cellref = sheet_obj.cell(row = data_col_loc[1]+1+trading_day,column = col_index+header_index)
                                    cellref.value = header_value_list[header_index]
                                    
        if(Commonapi.info == 1):
            logging.info('writeCPIOTMIOtoCPMISheet Ended')
                
#money inv calls getmoneyInvAllExpDateAll(symbol,contractype) which is in commonapi.py for each option type
#gets money inv dict for all existing expiration dates and calls local writeMaxPainDataToXl
        #input
        #symbol - stock symbol        
def money_inv(Symbol):
        if(Commonapi.info == 1):
            logging.info('money_inv Started')
        symbol = (str)(Symbol)
        strikeprice_list = []
        call_error,istockCSVfile =  dataapi.getStockCSVFile(0,symbol,ifilepath)
        if(call_error == globalheader.Success):    
                oxlfile = Commonapi.getoFile(Symbol,ofilename)    
                wb_obj = openpyxl.load_workbook(oxlfile)
                sheets = wb_obj.sheetnames
                
        #Loop around num of expiration x contracttype x getNumberColsData x strikeprices                    
                num_exp_dates = 0
                keys = []
                for contracttype in range(len(contractType)):
                    #call_error,money_inv_dict = Commonapi.getmoneyinv(Symbol,stockOptionExpDate,contracttype)
                    call_error,money_inv_dict,otm_dict,otmoi_dict,tm_dict = Commonapi.getmoneyInvAllExpDateAllData(Symbol,contracttype)
                    keys = money_inv_dict.keys()
                    num_exp_dates = len(keys)
                    if(call_error == globalheader.Success):
                        if(Commonapi.info == 1):
                            logging.info('%s %d','Commonapi.getmoneyInvAllExpDateAllData Success', call_error)
                        writeMoneyInvDataToXl(contracttype,wb_obj,money_inv_dict,Symbol)
                        writeTotalMoneyInvDatatoXl(contracttype,wb_obj,tm_dict,Symbol)
                        writeCPIOTMtoCPMISheet(contracttype,wb_obj,otm_dict,Symbol)
                        writeCPIOTMIOtoCPMISheet(contracttype,wb_obj,otmoi_dict,Symbol)
                    else:
                        logging.error('%s %d', 'Commonapi.getmoneyInvAllExpDateAllData Failed', call_error)
                        return call_error
                sheet_obj = wb_obj["All_Exp_CPMI"]                
                chartapi.insertTotalMoneyCPChart(sheet_obj,num_exp_dates,Symbol,keys,chart_col_loc,final_chart_insert_pos,length)

                chartapi.insertCPIOMchartToCPMISheet(sheet_obj,Symbol,keys,num_exp_dates,final_OTM_chart_insert_pos,chart_col_loc,length,data_col_loc)

                chartapi.insertCPIOMOIchartToCPMISheet(sheet_obj,Symbol,keys,num_exp_dates,final_OTMOI_chart_insert_pos,chart_col_loc,length,data_col_loc)
                wb_obj.save(oxlfile)
                if(Commonapi.info == 1):
                    logging.info('money inv Ended')
                return call_error

        else:
            logging.error('%s %d', 'money inv Failed', call_error)
            return call_error
        
if __name__ == "__main__":
        if(Commonapi.info == 1):
            logging.info('money inv  main Started')
        symbol = raw_input("Enter symbol :") 
        call_error = money_inv(symbol)
        if(call_error == globalheader.Success):
            if(Commonapi.info == 1):
                logging.info('%s %d','money_inv Success', call_error)
            else:
                logging.error('%s %d', 'money_inv  Failed', call_error)
        if(Commonapi.info == 1):
                logging.info('money inv  main Ended')
logging.warning('done')
