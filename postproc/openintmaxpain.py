## openintmaxpain.py calculates max pain of the stock of
## the given exp dates

import csv
import pandas as pd
from pandas import DataFrame
from datetime import date
from datetime import datetime
import time
import os
import sys
import os.path
import datetime
from os import path
import xlsxwriter
import openpyxl
from datetime import timedelta
import chartapi
import openintmi
import commonapi
csv.field_size_limit(sys.maxsize)

contractType =['CALL','PUT']
ifilepath = './../datacolls/output/'
ifilestocklist = './../datacolls/input/'+'StockList.csv'

ofilepath = './output/'
ofilename = ofilepath+'OImax_pain_mp'

maxpain_row_header = ["CallMP", "PutMP","TotalMP"]
expdate_chart_y_axis=['MAX_PAIN_CALL x 100','MAX_PAIN_PUT x 100 ','MAX_PAIN_TOTAL x 100']
expdate_chart_col_loc =['2','101','201']
expdate_chart_insert_pos =["A24","CV24","GR24"]
expdate_chart_insert_pos1 =["A40","CV40","GR40"]

debug = 0
OI_INDEX = 0
SP_INDEX = 1

#### Parsing the row cell data
def parseInputRowData(colFrom,data_frame,row):
    rowData = []
    rowData = data_frame.iloc[row,colFrom]
    if(rowData == 'U'):
        OI =0
        SP =0
        stockPrice = (float)(0)
        strikePrice = (float)(0)
    else:        
        rowData = str(rowData).replace('[','')
        rowData = str(rowData).replace(']','')
        rowData = str(rowData).split(':')
        OI = rowData[OI_INDEX]
        if((OI == 'U') or (OI == '') or (OI == '.')):
            OI = (int)(0)
        else:
            OI =(float)(rowData[OI_INDEX])
    return OI
    
#### Adding call/put max pain of individual strikeprice to excel sheet
def writeMaxPainDataToXl(contracttype,sheet_obj,strikeprices,strikeprices_max_pain,row,dateData):    
    if(contracttype == commonapi.Contracttype.CALL.value):
        if(row == 1):
            for i in range(len(maxpain_row_header)):
                col_index = (int)(expdate_chart_col_loc[i])-1
                cellref=sheet_obj.cell(row=1, column=col_index)
                cellref.value=maxpain_row_header[i]

                col_index = (int)(expdate_chart_col_loc[i])
                for item in range(len(strikeprices)):
                    cellref=sheet_obj.cell(row = 1, column=col_index+item)
                    cellref.value=strikeprices[item]
                
    current_row_index = row
    col_index= (int)(expdate_chart_col_loc[contracttype])-1 
    #print("current_row_index  after call is :",current_row_index)
    cellref=sheet_obj.cell(row = current_row_index+1, column=col_index)
    cellref.value= dateData
        
    col_index = (int)(expdate_chart_col_loc[contracttype])
    for item in range(0,len(strikeprices_max_pain)):                            
        cellref=sheet_obj.cell(row = current_row_index+1, column=col_index+item)
        cellref.value=strikeprices_max_pain[item]
        
#### Adding total call/put max pain of individual strikeprice to excel sheet
def writeTotalMaxPaintoXl(sheet_obj,strikeprices):
        max_rows = sheet_obj.max_row
        min_col = (int)(expdate_chart_col_loc[2])

        min_row = max_rows
        for row in range(2, max_rows+1):
            cellref=sheet_obj.cell(row = row, column=1)
            value = cellref.value
            cellref=sheet_obj.cell(row = row, column=min_col-1)
            cellref.value=value
            callmpvalue = []
            putmpvalue = []
            col_index = (int)(expdate_chart_col_loc[0])
            for item in range(len(strikeprices)):
                cellref=sheet_obj.cell(row = row, column=col_index+item)
                value = cellref.value
                callmpvalue.append(value)
            col_index = (int)(expdate_chart_col_loc[1])   
            for item in range(len(strikeprices)):
                cellref=sheet_obj.cell(row = row, column=col_index+item)
                value = cellref.value
                putmpvalue.append(value)
        
            for i in range(len(strikeprices)):
                cellref=sheet_obj.cell(row = row, column=min_col+i)
                total = callmpvalue[i]+putmpvalue[i]
                cellref.value=total 

#Adding each expiration to workbook
def addSheetToXl(wb_obj,symbol,sheetIndex,finalOptionExpDateList):
    sheet_name =  str(symbol)+finalOptionExpDateList
    if sheet_name in wb_obj.sheetnames:
        print("wb_obj.sheetnames is :",wb_obj.sheetnames)
                    
    if not sheet_name in wb_obj.sheetnames:
        wb_obj.create_sheet(index = sheetIndex , title = sheet_name)


def getStrikePriceMaxPain(strikepriceList,stockprice,strikePriceFromCSV,row,OptionExpDate,colFrom,data_frame,contracttype,max_pain):
    strikeexpdate = commonapi.getOptExpDateFromCSV(strikePriceFromCSV[row])
    strikeprice = strikepriceList[stockprice]

    if(strikeexpdate == OptionExpDate):
        strikePrice = commonapi.getOptionStrikePrice(strikePriceFromCSV[row])                        
        stockPrice = strikeprice
    
        stockPrice = (float)(stockPrice)
        strikePrice = (float)(strikePrice)
        OI = parseInputRowData(colFrom,data_frame,row)
        
        if(contracttype == commonapi.Contracttype.CALL.value):
            if(stockPrice > strikePrice):                                           
               maxpain = abs(stockPrice-strikePrice)
               max_pain +=maxpain*OI
            elif(stockPrice < strikePrice):
               max_pain += 0
            else:
               if(stockPrice == strikePrice):
                    max_pain += 0
               else:
                    max_pain += 0

        if(contracttype == commonapi.Contracttype.PUT.value):
            if(stockPrice < strikePrice):
               maxpain = abs(stockPrice-strikePrice)
               max_pain += maxpain*OI
            elif(stockPrice > strikePrice):
                max_pain += 0 
            else:
                if(stockPrice == strikePrice):
                    max_pain += 0
                else:
                    max_pain += 0
    return max_pain
                           
def maxpain(symbol,num_expiration,getNumberColsData):

    sheet_count = 0
    finalOptionExpDateList =[]
    print("getNumberColsData,num_expiration is :",getNumberColsData,num_expiration)
    
    oxlfile = commonapi.createoFile(symbol,ofilename)
    if os.path.exists(oxlfile):
        print("file exists:",oxlfile)
        os.remove(oxlfile)
        print("File Removed!")
    
    if os.path.exists(oxlfile):
        wb_obj = openpyxl.load_workbook(oxlfile)
        print("File Exists:",oxlfile)       
    else:
        print("File does not Exists :",oxlfile)
        wb_obj = openpyxl.Workbook()

    istockCSVfile =  commonapi.getStockCSVFile(0,symbol,ifilepath)
    data_frame = pd.read_csv(istockCSVfile, index_col = False)
    finalOptionExpDateList = commonapi.getExpDatesListFromCSV(data_frame,num_expiration)

#Loop around num of expiration x contracttype x getNumberColsData x strikeprices
    for expdate in range(0,len(finalOptionExpDateList)):
        colFrom = 0
        strikeprices = []
        strikeprices_max_pain = []
        
        for contracttype in range(len(contractType)):
            del strikeprices[:]
            del strikeprices_max_pain[:]
            data_frame = 0

            if(contracttype == commonapi.Contracttype['CALL'].value):
                addSheetToXl(wb_obj,symbol,expdate,finalOptionExpDateList[expdate])
                                
            sheets = wb_obj.sheetnames
            sheet_obj = wb_obj[sheets[expdate]]

            istockCSVfile =  commonapi.getStockCSVFile(contracttype,symbol,ifilepath)
         
            if os.path.exists(istockCSVfile):
                print("File Exists")
                data_frame,columnList,colFrom,strikePriceFromCSV = commonapi.getDataFrame(istockCSVfile,getNumberColsData)

                strikepriceList = []
                strikepriceList = commonapi.strikePricesFromCSV(finalOptionExpDateList[expdate],data_frame)
                strikeprices = strikepriceList
                
                for col in range(0,getNumberColsData):                    
                    del strikeprices_max_pain[:]
                    
                    for stockprice in range(0,len(strikepriceList)):
                        max_pain = 0
                        
                        for row in range(0,len(strikePriceFromCSV)):
                            max_pain = getStrikePriceMaxPain(strikepriceList,stockprice,strikePriceFromCSV,row,finalOptionExpDateList[expdate],colFrom,data_frame,contracttype,max_pain)                                  
                                                                                                                                                   
                        strikeprices_max_pain.append(max_pain)
                        
     #### Adding call max pain of individual strikeprice to excel sheet
                    row = col+1
                    writeMaxPainDataToXl(contracttype,sheet_obj,strikeprices,strikeprices_max_pain,row,columnList[colFrom])
                    colFrom += 1                                  

    #### Calculating Total Max Pain of call and put then adding data to excel sheet
        writeTotalMaxPaintoXl(sheet_obj,strikeprices)

    ### call chartapi to insert chart
        chartapi.insertLineChart(sheet_obj,strikeprices,symbol,finalOptionExpDateList[expdate],expdate_chart_y_axis,expdate_chart_col_loc,expdate_chart_insert_pos)
        
    wb_obj.save(oxlfile)
                           
if __name__ == "__main__":
    symbol = raw_input("Enter symbol :") 
    print(symbol) 
##    num_expiration = raw_input("Enter number of expiration dates data you want : ") 
##    print(num_expiration)
##    getNumberColsData = raw_input("Enter number of days data you want : ") 
##    print(getNumberColsData)
    num_expiration = 8
    getNumberColsData = 20
                           
    maxpain(symbol,num_expiration,getNumberColsData)

    #call openint money inv py 
    openintmi.money_inv(symbol,num_expiration,getNumberColsData)
                    
