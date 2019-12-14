import csv
import pandas as pd
from pandas import DataFrame
from datetime import date
from datetime import datetime
import time
import os
import os.path
import datetime
from os import path
import xlsxwriter
import openpyxl
from datetime import timedelta
from openpyxl.chart import BarChart,Reference,Series,LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.label import DataLabel

global CSVOptionSymbolList
import sys
######## working ti implement otm
endpoint = r"https://api.tdameritrade.com/v1/marketdata/chains"
stockSymbol = []
stockOptionType = []
stockCsvFiles = []
todayDate = []
expDate = []
getNumberColsData = 20
dates =[]
num_of_exp = 0

csv.field_size_limit(sys.maxsize)

contractType =['CALL','PUT']
ifilepath = './../datacolls/output/'
ifilestocklist = './../datacolls/input/'+'StockList.csv'

ofilepath = './output/'
ifile = ofilepath
ofilename = ofilepath+'OImax_pain_mp'

header = ["CallMI", "PUTM","TOTALMI"]
TTM =["CALLTM","PUTTM"]
final_chart_insert_pos = ["A20","K20","T20","AD20","AN20","AX20","BH20","BR20"]
final_chart_col_loc = ['2','12','22','32','42','52','62','72']

expdate_chart_insert_pos =["A20","BR20","EJ20","IF20","LH20","OJ20","SF20"]
expdate_chart_col_loc =['2','71','141','241','321','401','500']
expdate_chart_y_axis=['MAX_PAIN_CALL','MAX_PAIN_PUT','MAX_PAIN_TOTAL','MONEY_INV_CALL','MONEY_INV_PUT','MONEY_INV_TOTAL','TOTAL_CALL_PUT_TM']

OITM =['CallITM','CallOTM','PutITM','PutOTM']
final_OTM_chart_insert_pos = ["A70","K70","T70","AD70","AN70","AX70","BH70","BR70"]
final_OTM_chart_col_loc = ['2','12','22','32','42','52','62','72']
final_OTM_chart_y_axis =['CPIOTM x K']
num_of_chart_per_sheet = 7
num_headers = 4
debug = 0
OI_INDEX = 0
SP_INDEX = 1
OP_INDEX = 3


symbol = raw_input("Enter symbol :") 
print(symbol) 
num_expiration = raw_input("Enter number of expiration dates data you want : ") 
print(num_expiration) 

sym = symbol
totalStocks = 1

def getStockCSVFiles(index,stockIndex):       
    stockCsvfilename = ifilepath + symbol+'_'+contractType[index]+'.csv'
    return stockCsvfilename

def createoFile():
    today = datetime.date.today()
    today= datetime.datetime.strftime((today), '%m_%d_%Y')
    stockCsvfilename = ofilename+'_'+symbol+'_'+today+'.xlsx'
    return stockCsvfilename

def getOptExpDateFromCSV(optionSymbol):
    symbol,date = optionSymbol.split('_')
    date = list(str(date))
    length = len(date)
    expdate = '20'+date[4]+date[5]+date[0]+date[1]+date[2]+date[3]
    #print("expdate is :",expdate)
    return expdate


def removeDuplicateDates(duplicate): 
    final_list = [] 
    for num in duplicate: 
        if num not in final_list: 
            final_list.append(num) 
    return final_list

def getOptionStrikePrice(optionSymbol):
    n = 7
    symbol,strikeprice = optionSymbol.split('_')
    strikeprice  = list(str(strikeprice ))
    length = len(strikeprice )
    strikeprice= strikeprice[n:]
    strikeprice =''.join(strikeprice)
#       print("strikeprice is.................. :",strikeprice)
    return strikeprice

def mainloop():

    global CSVOptionSymbolList
    global totalStocks
    sheet_count = 0
    num_rows = 0

    ofile = createoFile()    
    wb_obj = openpyxl.load_workbook(ofile)
    sheets = wb_obj.sheetnames    
    depth = [[[]]]
    OTM = [[[]]]
    
    for index in range(totalStocks):
        Index = index
        rows = []        
        colCount = 0
        rowCount = 0

        for contracttype in range(0,1):
            ifile =  getStockCSVFiles(contracttype,Index)
            if os.path.exists(ifile):
                print("File Exists:",ofile)                                    
            else:
                print("File does not Exists:",ofile)
                 
            if os.path.exists(ifile):
                print("File Exists")
                data_frame = pd.read_csv(ifile, index_col = False)
                countRows = data_frame.shape[0]
                countCols = data_frame.shape[1]
                
                CSVOptionSymbolList= data_frame.Symbol.tolist()
                if(debug == 1):
                    print("CSVOptionSymbolList  and length is  :",CSVOptionSymbolList,len(CSVOptionSymbolList))
                
                expDateList =[]
                for index in range(0,len(CSVOptionSymbolList)):
                    expdatelist = getOptExpDateFromCSV(CSVOptionSymbolList[index])
                    expDateList.append(expdatelist)                
                    
                finalOptionExpDateList=[]
                finalOptionExpDateList = removeDuplicateDates(expDateList)
                finalOptionExpDateList= sorted(finalOptionExpDateList)

                print("finalOptionExpDateList is :",finalOptionExpDateList)
                
                
                finalOptionExpDateList1 =[]
                for i in range(0,(int)(num_expiration)):
                    finalOptionExpDateList1.append(finalOptionExpDateList[i])
                    expDate.append(finalOptionExpDateList[i])

                print("finalOptionExpDateList1 and len(finalOptionExpDateList1 ) is :",finalOptionExpDateList1,len(finalOptionExpDateList1))

            num_of_exp = len(finalOptionExpDateList1)
            for i in range(0,len(finalOptionExpDateList1)):
                depth.append([])
                for j in range(0,len(contractType)):
                    depth[i].append([])
                    for k in range(0,getNumberColsData):
                        depth[i][j].append([])
            for i in range(0,len(finalOptionExpDateList1)):
                for j in range(0,len(contractType)):
                    for k in range(0,getNumberColsData):
                        val = (int)(0)
                        depth[i][j][k] = val

            for i in range(0,len(finalOptionExpDateList1)):
                OTM.append([])                
                for j in range(0,getNumberColsData):
                    OTM[i].append([])
                    for k in range(0,4):
                        OTM[i][j].append([])
            for i in range(0,len(finalOptionExpDateList1)):
                for j in range(0,getNumberColsData):
                    for k in range(0,4):
                        val = (int)(0)
                        OTM[i][j][k] = val
            for expdate in range(0,len(finalOptionExpDateList1)):
                print("finalOptionExpDateList1 ;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;:",finalOptionExpDateList1[expdate])
                colFrom = 0
                strikeprices = []
                strike_money_inv = []
                col=[]
                
                for contracttype in range(len(contractType)):
                    del strikeprices[:]
                    del strike_money_inv[:]
                    m_col_call = 0
                    m_col_put =0
                    m_row_call = 0
                    m_row_put =0
                    if(expdate == 0):
                        if 'All_Exp_CPMI' in wb_obj.sheetnames:
                            print("wb_obj.sheetnames is :",wb_obj.sheetnames)
                            std=wb_obj.get_sheet_by_name('All_Exp_CPMI')
                            wb_obj.remove_sheet(std)
                        if not 'All_Exp_CPMI' in wb_obj.sheetnames:
                            wb_obj.create_sheet('All_Exp_CPMI')                   
                    sheet_obj = wb_obj[sheets[expdate]]
                    print(sheet_obj.max_column)
                    m_row = sheet_obj.max_row
                    num_rows = sheet_obj.max_row
                    print("number of rows and columns is :",m_row,sheet_obj.max_column)

                    symbollist =[]
                    for i in range(2, m_row + 1): 
                        cell_obj = sheet_obj.cell(row = i, column = 1) 
                        symbollist.append(cell_obj.value)
                    #print("symbollist is :",symbollist)
                        
                    ifile =  getStockCSVFiles(contracttype,Index)
                    if os.path.exists(ifile):
                        print("File Exists:",ifile)                                    
                    else:
                        print("File does not Exists:",ifile)
                 
                    if os.path.exists(ifile):
                        print("File Exists")
                        data_frame = pd.read_csv(ifile, index_col = False)
                        countRows = data_frame.shape[0]
                        countCols = data_frame.shape[1]
                        
                
                        CSVOptionSymbolList= data_frame.Symbol.tolist()
                        if(debug == 1):
                            print("CSVOptionSymbolList  and length is  :",CSVOptionSymbolList,len(CSVOptionSymbolList))
                        length = len(CSVOptionSymbolList)

                        strikepriceList =[]
                        for index in range(0,len(CSVOptionSymbolList)):                                    
                            symbolexpdate1 = getOptExpDateFromCSV(CSVOptionSymbolList[index])
                            if(symbolexpdate1 == finalOptionExpDateList1[expdate]):
                                strikeprice = getOptionStrikePrice(CSVOptionSymbolList[index])
                                strikepriceList.append(strikeprice)
                        strikepriceList = removeDuplicateDates(strikepriceList)
                        strikepriceList = sorted(strikepriceList)
                        print("strikepriceList = sorted(strikepriceList) is :",strikepriceList )
                        
                        rowdata =[]                            
                        columnList= data_frame.columns.tolist()
                        colFrom = countCols-getNumberColsData
                        colTo = countCols

                        print("colTo and colFrom is :",colTo,colFrom)
                        
                        for col in range(0,getNumberColsData):
                            
                            del strikeprices[:]
                            del strike_money_inv[:]
                            for stockprice in range(0,len(strikepriceList)):
                                ent = 0
                                money_inv = 0                                
                                for row in range(0,len(CSVOptionSymbolList)):
                                    symbolexpdate = (getOptExpDateFromCSV(CSVOptionSymbolList[row]))
                                    strikeprice = (float)(getOptionStrikePrice(CSVOptionSymbolList[row]))
                                    #strikeprice = strikepriceList[stockprice]
                                    strikeprice = (float)(strikeprice)
                                    if((symbolexpdate == finalOptionExpDateList1[expdate]) and ((float)(strikepriceList[stockprice]) == strikeprice)):
                                        #print("symbolexpdate == finalOptionExpDateList1[expdate])  ")
                                        ent +=1
                                        if(ent == 1):
                                            strikeprices.append(strikepriceList[stockprice])
                                        
                                        data_frame = pd.read_csv(ifile, index_col = False)
                                        countRows = data_frame.shape[0]
                                        countCols = data_frame.shape[1]                                        
                                        rowdata = data_frame.iloc[row,colFrom]
                                        if(rowdata == 'U'):
                                            OI =0
                                            SP =0
                                            OP = 0
                                            money_inv = OI*OP
                                        else:
                                            rowdata = str(rowdata).replace('[','')
                                            rowdata = str(rowdata).replace(']','')
                                            rowdata = str(rowdata).split(':')
                                            #print("data is :",rowdata)
                                            OIref =rowdata[OI_INDEX]
                                            if((OIref == 'U') or (OIref == '.') or (OIref == '')):
                                                OI = (int)(0)
                                            else:
                                                OI =(float)(rowdata[OI_INDEX])
                                            OPref =rowdata[OP_INDEX]
                                            if((OPref == 'U') or (OPref =='.') or (OPref == '')):
                                                OP = (int)(0)
                                            else:
                                                OP =(float)(rowdata[OP_INDEX])

                                            SPref =rowdata[SP_INDEX]
                                            if((SPref == 'U') or (SPref =='.') or (SPref == '')):
                                                SP = (int)(0)
                                            else:
                                                SP =(float)(rowdata[SP_INDEX])

                                            if(col == 19):
                                                print("sp is ......................................... :",SP)
                                            money_inv = OI*OP*100
                                            money_inv = money_inv/1000

                                            if(contracttype == 0):
                                                if(SP > strikeprice):
                                                    value = OTM[expdate][col][0]
                                                    value += money_inv 
                                                    OTM[expdate][col][0] = value
                                                else:
                                                    value = OTM[expdate][col][1]
                                                    value += money_inv 
                                                    OTM[expdate][col][1] = value

                                            if(contracttype == 1):
                                                if(SP < strikeprice):
                                                    value = OTM[expdate][col][2]
                                                    value += money_inv 
                                                    OTM[expdate][col][2] = value
                                                else:
                                                    value = OTM[expdate][col][3]
                                                    value += money_inv 
                                                    OTM[expdate][col][3] = value
                                            
                                           
                                strike_money_inv.append(money_inv)
                                #print("depth[item1][contracttype][i] ,expdate,contracttype,money_inv:",depth[expdate][contracttype][col],expdate,contracttype,money_inv)
                                                               
                            if(contracttype == 0):
                                m_col = (int)(expdate_chart_col_loc[3])-1
                                print(sheet_obj.max_column)
                                if(expdate == 0):
                                    cellref=sheet_obj.cell(row=1, column=m_col)
                                    cellref.value=header[0]
                                    
                                    for item in range(len(strikeprices)):
                                        cellref=sheet_obj.cell(row = 1, column=m_col+item+1)
                                        cellref.value=strikeprices[item]
                                m_row_call += 1
                                #m_row_call = (int)(expdate_chart_col_loc[3])
                                cellref=sheet_obj.cell(row =m_row_call+1, column=m_col)
                                cellref.value=columnList[colFrom]
                                col_index = (int)(expdate_chart_col_loc[3])
                                value = 0
                                for item3 in range(0,len(strike_money_inv)):                            
                                    cellref=sheet_obj.cell(row = m_row_call+1, column=col_index+item3)
                                    cellref.value=strike_money_inv[item3]
                                    value += strike_money_inv[item3]
                                depth[expdate][contracttype][col] = value
                                #print("depth[expdate][contracttype][col], value and contracttype is :",depth[expdate][contracttype][col],value,contracttype)


                            if(contracttype == 1):
                                m_col = (int)(expdate_chart_col_loc[4])-1
                                
                                print(sheet_obj.max_column)
                                if(expdate == 0):
                                    cellref=sheet_obj.cell(row=1, column=m_col)
                                    cellref.value=header[1]                                 
                                    for item in range(len(strikeprices)):
                                        cellref=sheet_obj.cell(row = 1, column=m_col+item+1)
                                        cellref.value=strikeprices[item]                                        
                                m_row_put += 1
                                cellref=sheet_obj.cell(row = m_row_put+1, column=m_col)
                                cellref.value=columnList[colFrom]
                                col_index = (int)(expdate_chart_col_loc[4])
                                value = 0
                                for item3 in range(0,len(strike_money_inv)):                            
                                    cellref=sheet_obj.cell(row = m_row_put+1, column=col_index+item3)
                                    cellref.value=strike_money_inv[item3]                                    
                                    value += strike_money_inv[item3]
                                depth[expdate][contracttype][col] = value
                                #print("depth[expdate][contracttype][col], value and contracttype is :",depth[expdate][contracttype][col],value,contracttype)
                            colFrom += 1
                
            
                m_rows = sheet_obj.max_row            
                m_col = (int)(expdate_chart_col_loc[5])-1
                print("m_col is :",m_col)
                cellref=sheet_obj.cell(row=1, column=m_col)
                cellref.value=header[2]
                
                tmcol_index = (int)(expdate_chart_col_loc[6])
                cellref = sheet_obj.cell(row =1,column =tmcol_index)
                cellref.value = TTM[0]


                cellref = sheet_obj.cell(row =1,column =tmcol_index+1)
                cellref.value = TTM[1]

                for item in range(len(strikeprices)):
                    cellref=sheet_obj.cell(row = 1, column=m_col+item+1)
                    cellref.value=strikeprices[item]

                for i in range(2, m_rows+1):
                    cellref=sheet_obj.cell(row = i, column=(int)(expdate_chart_col_loc[3])-1)
                    value = cellref.value
                    cellref=sheet_obj.cell(row = i, column=m_col)
                    cellref.value=value
                    if(expdate == 0):
                        dates.append(value)
                                            
                    callvalue = []
                    putvalue = []
                    callstriketm = 0
                    putstriketm = 0
                    col_index = (int)(expdate_chart_col_loc[3])
                    for item in range(len(strikeprices)):
                        cellref=sheet_obj.cell(row = i, column=col_index+item)
                        value = cellref.value
                        callvalue.append(value)
                        callstriketm += value
                        
                    col_index = (int)(expdate_chart_col_loc[4])   
                    for item in range(len(strikeprices)):
                        cellref=sheet_obj.cell(row = i, column=col_index+item)
                        value = cellref.value
                        putvalue.append(value)
                        putstriketm += value
                    
                    for item in range(len(strikeprices)):
                        cellref=sheet_obj.cell(row = i, column=m_col+item+1)                        
                        total = callvalue[item]+putvalue[item]
                        cellref.value=total

                    cellref = sheet_obj.cell(row = i,column = tmcol_index)
                    cellref.value = callstriketm
                    cellref = sheet_obj.cell(row =i,column = tmcol_index+1)
                    cellref.value = putstriketm

                for j in range(0,num_of_chart_per_sheet-1):
                    print("chart :",num_of_chart_per_sheet+1)
                    min_cols = (int)(expdate_chart_col_loc[j])
                    max_cols = len(strikeprices)+min_cols
                    m_rows = sheet_obj.max_row

                    chart = LineChart()
                    chart.type = "line"
                    chart.title = symbol+'_'+finalOptionExpDateList1[expdate]
                    chart.y_axis.title = expdate_chart_y_axis[j]
                    chart.x_axis.title = 'Dates'
                    data = Reference(sheet_obj, min_col=min_cols, min_row=1, max_row=m_rows, max_col=max_cols)
                    cats = Reference(sheet_obj, min_col=(int)(expdate_chart_col_loc[0])-1, min_row=2, max_row=m_rows)
                    chart.add_data(data, titles_from_data=True)
                    chart.dataLabels = DataLabelList()
                    chart.dataLabels.showVal = True
                    chart.set_categories(cats)
                    sheet_obj.add_chart(chart, expdate_chart_insert_pos[j])

                print("chart :",num_of_chart_per_sheet)
                min_cols = tmcol_index
                max_cols = 1+min_cols
                m_rows = sheet_obj.max_row

                chart = BarChart()
                chart.type = "col"
                chart.grouping = "stacked"
                chart.overlap = 100
                chart.title = symbol+'_'+finalOptionExpDateList1[expdate]
                chart.y_axis.title = expdate_chart_y_axis[6]
                chart.x_axis.title = 'Dates'
                data = Reference(sheet_obj, min_col=min_cols, min_row=1, max_row=m_rows, max_col=max_cols)
                cats = Reference(sheet_obj, min_col=1, min_row=2, max_row=m_rows)
                chart.add_data(data, titles_from_data=True)
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showVal = True
                chart.set_categories(cats)
                sheet_obj.add_chart(chart, expdate_chart_insert_pos[6])
                
    
    sheet_obj = wb_obj["All_Exp_CPMI"]
    cellref = sheet_obj.cell(row = 1,column =1)
    cellref.value = "MICP"
    
    for item in range(len(dates)):
        cellref = sheet_obj.cell(row = item+2,column = 1)
        cellref.value = dates[item]
    cols = 2    
    for i in range(0,num_of_exp):
        min_cols = (int)(final_chart_col_loc[i])
        max_cols =min_cols+1
        cellref = sheet_obj.cell(row = 1,column =min_cols)
        cellref.value = "CALLTM"
        cellref = sheet_obj.cell(row = 1,column =max_cols)
        cellref.value = "PUTTM"
        m_cols = (int)(final_chart_col_loc[i])
        for j in range(0,2):            
            for k in range(0,getNumberColsData): 
                cellref=sheet_obj.cell(row = k+2, column=m_cols)                        
                total = depth[i][j][k]
                #print("depth[i][j][k] is :",depth[i][j][k])
                cellref.value=(int)(total)
            m_cols += 1
    
    for i in range(0,num_of_exp):
        print("chart :",i+1)
        min_cols = (int)(final_chart_col_loc[i])
        max_cols =min_cols+1
        m_rows = sheet_obj.max_row
        print("m_rows is :",m_rows)
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = symbol+'_'+expDate[i]
        chart.y_axis.title = 'TOTAL_CALL_PUT_TM x K'
        chart.x_axis.title = 'Dates'
        data = Reference(sheet_obj, min_col=min_cols, min_row=1, max_row=m_rows, max_col=max_cols)
        cats = Reference(sheet_obj, min_col=1, min_row=2, max_row=m_rows)
        chart.add_data(data, titles_from_data=True)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.set_categories(cats)
        sheet_obj.add_chart(chart, final_chart_insert_pos[i])


    min_rows = 42
    cellref = sheet_obj.cell(row = min_rows,column =1)
    cellref.value = "Dates"
    
    for item in range(len(dates)):
        cellref = sheet_obj.cell(row = item+min_rows+1,column = 1)
        cellref.value = dates[item]

    for i in range(0,num_of_exp):
        min_cols = (int)(final_OTM_chart_col_loc[i])
        max_cols =min_cols+3
        for j in range(0,4):
            cellref = sheet_obj.cell(row= min_rows,column =min_cols+j)
            cellref.value = OITM[j]
        m_cols = (int)(final_OTM_chart_col_loc[i])   
        for row in range(0,getNumberColsData):               
            for col in range(0,4):
                #print(",i,col,row is OTM[i][row][col]  :",i,col,row,OTM[i][row][col])
                cellref=sheet_obj.cell(row = row+min_rows+1, column=m_cols+col)
                total = OTM[i][row][col]
                #print("depth[i][j][k] is :",OTM[i][row][col])
                cellref.value=(int)(total)
        
    
    for i in range(0,num_of_exp):
        print("chart :",i+1)
        min_cols = (int)(final_OTM_chart_col_loc[i])
        max_cols =min_cols+3
        max_rows = min_rows+getNumberColsData
        print("m_rows is :",m_rows)
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = symbol+'_'+expDate[i]
        chart.y_axis.title = final_OTM_chart_y_axis[0]
        chart.x_axis.title = 'Dates'
        data = Reference(sheet_obj, min_col=min_cols, min_row=min_rows, max_row=max_rows, max_col=max_cols)
        cats = Reference(sheet_obj, min_col=1, min_row=min_rows+1, max_row=max_rows)
        chart.add_data(data, titles_from_data=True)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.set_categories(cats)
        sheet_obj.add_chart(chart, final_OTM_chart_insert_pos[i]) 
    wb_obj.save(ofile)
mainloop() 
                    
print('done')
