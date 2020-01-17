
import csv
import pandas as pd
from pandas import DataFrame
from datetime import date
from datetime import datetime
import time
import os
import os.path
import datetime
from os import path
import openpyxl
from datetime import timedelta
from openpyxl.chart import BarChart,Reference,Series,LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.label import DataLabel
from collections import OrderedDict
import sys
import chartapi
import Commonapi
import dataapi
import globalheader

csv.field_size_limit(sys.maxsize)

contractType =['CALL','PUT']
ifilepath = './../datacolls/output/'

ofilepath = './output/'
ifile = ofilepath
ofilename = ofilepath+'OImax_pain_ExpDate'
#thses datas are used by chart insertion and data insertion of the money invested of call and put

money_inv_row_header = ["CallMIxK", "PutMIxK"]
TTM =["CallTM","PutTM"]
expdate_chart_y_axis=['MONEY_INV_CALL x K','MONEY_INV_PUT x K']
final_chart_col_loc = ['2','4']
expdate_chart_col_loc =['301','451']
final_TM_chart_col_loc = ['2','3']
expdate_chart_insert_pos =["KN64","OJ64"]
data_col_loc = []
chart_insert_loc=[]
chart_col_loc = ['2']
length = 0
final_chart_insert_pos = ["A60"]
final_OTM_chart_insert_pos = ["A145"]
final_OTMOI_chart_insert_pos = ["A230"]

#Adding last sheet "ALL_MICP" to workbook which contain data like OTM,ITM,total money inv of call/put of requested expiration dates
#addSheetToXl(wb_obj,symbol):
        #input
        #wb_obj - xl file
        #symbol - stock expiration symboldef addSheetToXl(wb_obj,symbol):
    if(Commonapi.info == 1):
        logging.info('addSheetToXl Started')
    if symbol in wb_obj.sheetnames:
        if(Commonapi.debug == 1):
            logging.debug('%s %s', 'sheet_name present in wb_obj.sheetnames is', symbol,wb_obj.sheetnames) 
        std=wb_obj.get_sheet_by_name(symbol)
        wb_obj.remove_sheet(std)
    if not symbol in wb_obj.sheetnames:
        if(Commonapi.debug == 1):
            logging.debug('%s %s', 'sheet_name not present in wb_obj.sheetnames is', symbol,wb_obj.sheetnames)
        wb_obj.create_sheet(symbol)
    if(Commonapi.info == 1):
        logging.info('addSheetToXl Ended')

# Adding call/put max pain data of individual strikeprice of all expiration dates of all trading days to excel sheet
# calls insertLineChart which is in chartapi.py 
#writeMoneyInvDataToXl(contracttype,wb_obj,money_inv_dict,Symbol,stockOptionExpDate):
        #input
        #contracttype - call or put
        #wb_obj - xl file
        # money_inv_dict - money inv dict data
        #symbol - stock expiration symbol
        #stockOptionExpDate - stock exp date
def writeMoneyInvDataToXl(contracttype,wb_obj,money_inv_dict,Symbol,stockOptionExpDate):
        if(Commonapi.info == 1):
            logging.info('writeMoneyInvDataToXl Started')
        trading_days_list = money_inv_dict.keys()
        sheet_obj_ref = 0
        strikeprice_list = []
        
        if(contracttype == Commonapi.Contracttype['CALL'].value):
                sym = 'All_Exp_CPMI'
                addSheetToXl(wb_obj,sym)
        sheets = wb_obj.sheetnames
        sheet_obj = wb_obj[sheets[0]]
        sheet_obj_ref = sheet_obj

        for expdatekey in range(0,len(trading_days_list)):
                
                strikeprice_list = money_inv_dict[trading_days_list[expdatekey]].keys()
                strikeprice_value_list = money_inv_dict[trading_days_list[expdatekey]].values()
                
                if(contracttype == Commonapi.Contracttype.CALL.value):
                        col_index = (int)(expdate_chart_col_loc[contracttype])
                        cellref=sheet_obj.cell(row=1, column=col_index-1)
                        cellref.value=money_inv_row_header[contracttype]
                        cellref=sheet_obj.cell(row=expdatekey+2, column=col_index-1)
                        cellref.value=trading_days_list[expdatekey]

                        col_index = (int)(expdate_chart_col_loc[contracttype])
                        for i in range(len(strikeprice_list)):
                                cellref=sheet_obj.cell(row=1, column=col_index+i)
                                cellref.value=strikeprice_list[i]
                        for item in range(len(strikeprice_list)):
                                cellref=sheet_obj.cell(row = expdatekey+2, column=col_index+item)
                                cellref.value=strikeprice_value_list[item]

                else:
                        col_index = (int)(expdate_chart_col_loc[contracttype])
                        cellref=sheet_obj.cell(row=1, column=col_index-1)
                        cellref.value=money_inv_row_header[contracttype]
                        cellref=sheet_obj.cell(row=expdatekey+2, column=col_index-1)
                        cellref.value=trading_days_list[expdatekey]
                                        
                        for i in range(len(strikeprice_list)):
                                cellref=sheet_obj.cell(row=1, column=col_index+i)
                                cellref.value=strikeprice_list[i]

                        for item in range(len(strikeprice_list)):
                            cellref=sheet_obj.cell(row = expdatekey+2, column=col_index+item)
                            cellref.value=strikeprice_value_list[item]
        y_axis_label =[]
        y_axis_label.append(expdate_chart_y_axis[contracttype])
        chart_insert_pos = []
        chart_insert_pos.append(expdate_chart_insert_pos[contracttype])
        chart_col_loc=[]
        chart_col_loc.append(expdate_chart_col_loc[contracttype])
        chartapi.insertLineChart(sheet_obj_ref,strikeprice_list,Symbol,stockOptionExpDate,y_axis_label,chart_col_loc,chart_insert_pos,1)

        if(Commonapi.info == 1):
            logging.info('writeMoneyInvDataToXl Ended')
                
# Adding call/put OTM,ITM money inv data  to last sheet "ALL_MICP of individual strikeprice of a expiration dates of all trading days to excel sheet
# calls insertLineChart which is in chartapi.py 
#writeTotalMoneyInvDatatoXl(contracttype,wb_obj,otm_dict,Symbol,stockOptionExpDate):
        #input
        #contracttype - call or put
        #wb_obj - xl file
        # otm_dict - OTM,ITM money inv dict data
        #symbol - stock expiration symbol
        #stockOptionExpDate - stock option exp date
def writeCPIOTMtoCPMISheet(contracttype,wb_obj,otm_dict,Symbol,stockOptionExpDate):
        if(Commonapi.info == 1):
            logging.info('writeCPIOTMtoCPMISheet Started')
        sheet_obj = wb_obj["All_Exp_CPMI"]
        min_rows = sheet_obj.max_row
        min_cols = sheet_obj.max_column

        rows = min_rows+30 
                
        trading_days_list = otm_dict.keys()
        for expdatekey in range(0,len(trading_days_list)):
                if((contracttype == Commonapi.Contracttype.CALL.value) and (expdatekey == 0)):
                    data_col_loc.append(rows)
                    chart_insert_loc.append(rows+len(trading_days_list)+2)
                    cellref = sheet_obj.cell(row = data_col_loc[0],column =1)
                    cellref.value = str(Symbol)+'_'+str(stockOptionExpDate)
                    
                
                header_list = otm_dict[trading_days_list[expdatekey]].keys()
                header_value_list = otm_dict[trading_days_list[expdatekey]].values()

                if(contracttype == Commonapi.Contracttype.CALL.value):
                        col_index = (int)(final_chart_col_loc[contracttype])
                        cellref=sheet_obj.cell(row=data_col_loc[0]+1+expdatekey, column=1)
                        cellref.value=trading_days_list[expdatekey]

                        for header_index in range(0,len(header_list)):
                                cellref = sheet_obj.cell(row = data_col_loc[0],column = col_index+header_index)
                                cellref.value = header_list[header_index]

                        for header_index in range(0,len(header_list)):
                                cellref = sheet_obj.cell(row = data_col_loc[0]+1+expdatekey,column = col_index+header_index)
                                cellref.value = header_value_list[header_index]
                                                                                
                else:
                        
                        col_index = (int)(final_chart_col_loc[contracttype])
                        cellref=sheet_obj.cell(row=data_col_loc[0]+1+expdatekey, column=1)
                        cellref.value=trading_days_list[expdatekey]

                        for header_index in range(0,len(header_list)):
                                cellref = sheet_obj.cell(row = data_col_loc[0],column = col_index+header_index)
                                cellref.value = header_list[header_index]
                        min_rows = sheet_obj.max_row
                        min_cols = sheet_obj.max_column
                        for header_index in range(0,len(header_list)):
                                cellref = sheet_obj.cell(row = data_col_loc[0]+1+expdatekey,column = col_index+header_index)
                                cellref.value = header_value_list[header_index]
        if(Commonapi.info == 1):
            logging.info('writeCPIOTMtoCPMISheet Ended')
        
# Adding call/put total money data of individual strikeprice of all expiration dates of a trading days to excel sheet
# calls insertLineChart which is in chartapi.py 
#writeTotalMoneyInvDatatoXl(contracttype,wb_obj,tm_dict,Symbol,stockOptionExpDate):
        #input
        #contracttype - call or put
        #wb_obj - xl file
        # tm_dict - total money dict data
        #symbol - stock expiration symbol
        #stockOptionExpDate - stock option exp date
def writeTotalMoneyInvDatatoXl(contracttype,wb_obj,tm_dict,Symbol,stockOptionExpDate):
        if(Commonapi.info == 1):
            logging.info('writeTotalMoneyInvDatatoXl Started')
        global length
        sheet_obj = wb_obj["All_Exp_CPMI"]
        min_rows = sheet_obj.max_row
        min_cols = sheet_obj.max_column
        
        cellref = sheet_obj.cell(row = 1,column =1)
        cellref.value = str(Symbol)+'_'+str(stockOptionExpDate)
        min_rows = sheet_obj.max_row
        trading_days_list = tm_dict.keys()
        for expdatekey in range(0,len(trading_days_list)):    

                header_list = tm_dict[trading_days_list[expdatekey]].keys()
                header_value_list = tm_dict[trading_days_list[expdatekey]].values()
                if((contracttype == Commonapi.Contracttype.CALL.value) and (expdatekey == 0)):
                    chart_insert_loc.append(len(trading_days_list)+1+2)
                    length = len(trading_days_list)

                if(contracttype == Commonapi.Contracttype.CALL.value):
                        col_index = (int)(final_TM_chart_col_loc[contracttype])
                        cellref=sheet_obj.cell(row=expdatekey+2, column=col_index-1)
                        cellref.value=trading_days_list[expdatekey]

                        for header_index in range(0,len(header_list)):
                                cellref = sheet_obj.cell(row = 1,column = col_index+header_index)
                                cellref.value = header_list[header_index]

                        for header_index in range(0,len(header_list)):
                                cellref = sheet_obj.cell(row = expdatekey+2,column = col_index+header_index)
                                cellref.value = header_value_list[header_index]
                                                                                
                else:
                        col_index = (int)(final_TM_chart_col_loc[contracttype])
                        cellref=sheet_obj.cell(row=expdatekey+2, column=1)
                        cellref.value=trading_days_list[expdatekey]

                        for header_index in range(0,len(header_list)):
                                cellref = sheet_obj.cell(row = 1,column = col_index+header_index)
                                cellref.value = header_list[header_index]

                        for header_index in range(0,len(header_list)):
                                cellref = sheet_obj.cell(row = expdatekey+2,column = col_index+header_index)
                                cellref.value = header_value_list[header_index]
        if(Commonapi.info == 1):
            logging.info('writeTotalMoneyInvDatatoXl Ended')

# Adding call/put in the money and out-of-money  open  interest of individual strikeprice of a expiration dates of all trading days to excel sheet
# calls insertLineChart which is in chartapi.py 
#writeCPIOTMIOtoCPMISheet(contracttype,wb_obj,otmoi_dict,Symbol,stockOptionExpDate):
        #input
        #contracttype - call or put
        #wb_obj - xl file
        #otmoi_dict - money inv in-the-money and out-of-the-money open  interest dict data
        #symbol - stock expiration symbol
        #stockOptionExpDate - stock option exp date
def writeCPIOTMIOtoCPMISheet(contracttype,wb_obj,otmoi_dict,Symbol,stockOptionExpDate):
    if(Commonapi.info == 1):
        logging.info('writeCPIOTMIOtoCPMISheet Started')
        
    sheet_obj = wb_obj["All_Exp_CPMI"]
    min_rows = sheet_obj.max_row
    min_cols = sheet_obj.max_column

    rows = min_rows+30 
            
    trading_days_list = otmoi_dict.keys()
    for expdatekey in range(0,len(trading_days_list)):
            if((contracttype == Commonapi.Contracttype.CALL.value) and (expdatekey == 0)):
                data_col_loc.append(rows)
                chart_insert_loc.append(rows+len(trading_days_list)+1)
                cellref = sheet_obj.cell(row =  data_col_loc[1],column =1)
                cellref.value = str(Symbol)+'_'+str(stockOptionExpDate)
                            
            header_list = otmoi_dict[trading_days_list[expdatekey]].keys()
            header_value_list = otmoi_dict[trading_days_list[expdatekey]].values()

            if(contracttype == Commonapi.Contracttype.CALL.value):
                    col_index = (int)(final_chart_col_loc[contracttype])
                    cellref=sheet_obj.cell(row= data_col_loc[1]+1+expdatekey, column=1)
                    cellref.value=trading_days_list[expdatekey]
                    
                    for header_index in range(0,len(header_list)):
                            cellref = sheet_obj.cell(row =  data_col_loc[1],column = col_index+header_index)
                            cellref.value = header_list[header_index]
                            
                    for header_index in range(0,len(header_list)):
                            cellref = sheet_obj.cell(row =  data_col_loc[1]+1+expdatekey,column = col_index+header_index)
                            cellref.value = header_value_list[header_index]
                                                                            
            else:
                    col_index = (int)(final_chart_col_loc[contracttype])
                    cellref=sheet_obj.cell(row=data_col_loc[1]+1+expdatekey, column=1)
                    cellref.value=trading_days_list[expdatekey]
                    
                    for header_index in range(0,len(header_list)):
                            cellref = sheet_obj.cell(row = data_col_loc[1],column = col_index+header_index)
                            cellref.value = header_list[header_index]

                    for header_index in range(0,len(header_list)):
                            cellref = sheet_obj.cell(row = data_col_loc[1]+1+expdatekey,column = col_index+header_index)
                            cellref.value = header_value_list[header_index]

    if(Commonapi.info == 1):
        logging.info('writeCPIOTMIOtoCPMISheet Ended')

            
#money inv calls getmoneyInvAllExpDateAll(symbol,contractype) which is in commonapi.py for each option type
#gets money inv dict for a existing expiration dates and calls local writeMaxPainDataToXl
        #input
        #symbol - stock symbol
        #StockOptionExpDate - stock option exp date
def money_inv(Symbol,StockOptionExpDate):
        if(Commonapi.info == 1):
            logging.info('money_inv Started')
        symbol = (str)(Symbol)
        stockOptionExpDate =str(StockOptionExpDate)
        strikeprice_list = []
        call_error,istockCSVfile =  dataapi.getStockCSVFile(0,symbol,ifilepath)
        if(call_error == globalheader.Success):    
                oxlfile = Commonapi.getoFile(Symbol,ofilename)    
                wb_obj = openpyxl.load_workbook(oxlfile)
                sheets = wb_obj.sheetnames
                sheet_obj = wb_obj[sheets[0]]
                sheet_obj_ref = sheet_obj
        #Loop around num of expiration x contracttype x getNumberColsData x strikeprices                    
                num_exp_dates = 0
                keys = []
                for contracttype in range(len(contractType)):
                    #call_error,money_inv_dict = Commonapi.getmoneyinv(Symbol,stockOptionExpDate,contracttype)
                    call_error,money_inv_dict,otm_dict,otmoi_dict,tm_dict = Commonapi.getmoneyInvExpDateAllData(Symbol,stockOptionExpDate,contracttype)
                    keys = money_inv_dict.keys()
                    num_exp_dates = len(keys)
                    if(call_error == globalheader.Success):
                        if(Commonapi.info == 1):
                            logging.info('%s %d','Commonapi.getmoneyInvExpDateAllData Success', call_error)
                        writeMoneyInvDataToXl(contracttype,wb_obj,money_inv_dict,Symbol,stockOptionExpDate)
                        writeTotalMoneyInvDatatoXl(contracttype,wb_obj,tm_dict,Symbol,stockOptionExpDate)
                        writeCPIOTMtoCPMISheet(contracttype,wb_obj,otm_dict,Symbol,stockOptionExpDate)
                        writeCPIOTMIOtoCPMISheet(contracttype,wb_obj,otmoi_dict,Symbol,stockOptionExpDate)
                    else:
                        logging.error('%s %d', 'Commonapi.getmoneyInvExpDateAllData Failed', call_error)
                        return call_error
                sheet_obj = wb_obj["All_Exp_CPMI"]
                stockOptionExpDateList = []
                stockOptionExpDateList.append(stockOptionExpDate)
                chartapi.insertTotalMoneyCPChart(sheet_obj,1,Symbol,stockOptionExpDateList,chart_col_loc,final_chart_insert_pos,length)

                chartapi.insertCPIOMchartToCPMISheet(sheet_obj,Symbol,stockOptionExpDateList,1,final_OTM_chart_insert_pos,chart_col_loc,length,data_col_loc)

                chartapi.insertCPIOMOIchartToCPMISheet(sheet_obj,Symbol,stockOptionExpDateList,1,final_OTMOI_chart_insert_pos,chart_col_loc,length,data_col_loc)

                wb_obj.save(oxlfile)
                if(Commonapi.info == 1):
                        logging.info('money inv Ended')
                return call_error

        else:
            logging.error('%s %d', 'money inv Failed', call_error)
            return call_error
                
        
if __name__ == "__main__":
        if(Commonapi.info == 1):
                logging.info('money inv  main Started')
        symbol = raw_input("Enter symbol :") 
        StockOptionExpDate = raw_input("Enter stock exp date in format 'YYMMDD' :",)                                    
        call_error = money_inv(symbol,StockOptionExpDate)
        if(call_error == globalheader.Success):
                if(Commonapi.info == 1):
                    logging.info('%s %d','money_inv Success', call_error)
                else:
                    logging.error('%s %d', 'money_inv  Failed', call_error)
        if(Commonapi.info == 1):
                logging.info('money inv  main Ended')
logging.warning('done')
